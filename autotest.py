r"""
autotest.py - 범용 Python 단위 테스트 자동화 파이프라인 (Ollama 전용)
=======================================================================
이 파일은 수정하지 마세요. 프로젝트 설정은 fixtures.py 에만 작성합니다.

사전 준비:
    pip install requests openpyxl
    ollama pull qwen3-coder:30b

사용법 (Windows):
    python autotest.py --summary "로그인 기능 추가" --files config/config.py logic/LogicProcess.py
    python autotest.py --summary "..." --model qwen3-coder:8b --files ...

프로젝트별 설정:
    fixtures.py 파일을 autotest.py 와 같은 폴더에 작성하세요.
    fixtures.py 가 없으면 기본 픽스처만 제공됩니다.

출력 (output/ 폴더):
    generated_tests.py    pytest 코드 (인프라 autotest.py 작성 + 로직 Ollama 생성)
    test_result_report.md 실행 결과 리포트 (버그 분석 포함)
    test_result.json      CI/CD 연동용 JSON
    UI_테스트목록.xlsx     UI 수동 테스트 목록
=======================================================================
"""

import os, sys, ast, json, re, textwrap, argparse, subprocess
from pathlib import Path
from datetime import datetime
import requests

DEFAULT_MODEL = "qwen3-coder:30b"
DEFAULT_HOST  = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
OUTPUT_DIR    = Path("output")

# -----------------------------------------------------------------------
# 핵심 설계 원칙:
#   autotest.py  : 절대 수정하지 않는 범용 도구
#   fixtures.py  : 프로젝트마다 1회 작성하는 설정 파일
#   Ollama       : 테스트 로직(body)만 생성, 경로/import 생성 안 함
# -----------------------------------------------------------------------

SYSTEM_PROMPT = """\
당신은 Python 시니어 QA 엔지니어입니다.
개발자가 수정한 Python 코드를 분석하여 두 가지를 JSON으로 생성합니다.

1. test_classes : pytest 테스트 클래스 목록 (순수 테스트 로직만, import/경로/Mock 설정 제외)
2. ui_test_cases: UI 수동 테스트 케이스 목록

반드시 아래 JSON 구조만 출력하세요. 다른 텍스트는 절대 포함하지 마세요.
{
  "test_classes": [
    {
      "class_name": "TestSignalType",
      "description": "SignalType Enum 값 검증",
      "methods": [
        {
          "name": "test_exit_program_value",
          "docstring": "EXIT_PROGRAM 값이 -1인지 확인",
          "body": "assert cfg.SignalType.EXIT_PROGRAM.value == -1"
        }
      ]
    }
  ],
  "ui_test_cases": [
    {
      "no": 1,
      "category": "기능 구분",
      "title": "테스트 항목명",
      "precondition": "사전 조건",
      "steps": "1. 단계1\\n2. 단계2",
      "expected": "기대 결과",
      "priority": "P1"
    }
  ]
}

테스트 메서드 작성 규칙:
- body 에는 테스트 로직만 작성. import/경로/Mock 설정 코드 절대 포함 금지.
- fixtures.py 에 정의된 헬퍼 함수를 활용하면 됩니다.
- 일반적으로 사용 가능한 변수: cfg (config 모듈), logic (LogicProcess 인스턴스), q_out (Queue), ST (SignalType)
- 비동기 결과: result = q_out.get(timeout=3)
- 예외 테스트: with pytest.raises(SystemExit): ...
- docstring 은 한국어로 테스트 목적 설명

UI 테스트 케이스: 코드로 검증 불가능한 UI/사용자 시나리오만 작성
priority: P1(필수), P2(중요), P3(선택)
"""

USER_PROMPT_TEMPLATE = """\
## 수정 요약
{summary}

## 코드 구조
{code_structure}

## 수정된 파일 내용
{files_content}

위 코드를 분석하여 JSON을 생성하세요.
test_classes 의 body 는 테스트 로직만 작성하고, import/경로/Mock 코드는 절대 포함하지 마세요.
"""

# -----------------------------------------------------------------------
# fixtures.py 로드
# -----------------------------------------------------------------------


def _extract_var(code: str, var_name: str, default):
    """fixtures.py 에서 특정 변수의 값을 파싱해서 반환한다."""
    import ast as _ast
    for line in code.splitlines():
        line = line.strip()
        if line.startswith(var_name + " ") or line.startswith(var_name + "="):
            try:
                val = _ast.literal_eval(line.split("=", 1)[1].strip())
                return val
            except Exception:
                pass
    return default


def load_fixtures(project_root: Path) -> str:
    """
    fixtures.py 파일을 읽어 문자열로 반환한다.
    없으면 기본 픽스처(최소한의 _load_module 헬퍼)를 반환한다.
    """
    fixtures_path = project_root / "fixtures.py"
    if fixtures_path.exists():
        print(f"  ✓ fixtures.py 로드: {fixtures_path}")
        return fixtures_path.read_text(encoding="utf-8")
    else:
        print(f"  ℹ  fixtures.py 없음 — 기본 픽스처 사용")
        print(f"     (더 정밀한 테스트를 원하면 fixtures.py 를 작성하세요)")
        return _default_fixtures()


def _default_fixtures() -> str:
    """fixtures.py 가 없을 때 사용하는 최소한의 기본 픽스처"""
    return '''\
# 기본 픽스처 (autotest.py 자동 생성)
# fixtures.py 를 작성하면 MOCK_MODULES, _load_config, _make_logic 등을 커스터마이징할 수 있습니다.

import importlib.util as _ilu

# Mock 처리할 패키지 목록 (기본값: 비어 있음 — fixtures.py 에서 재정의 가능)
MOCK_MODULES = []

def _load_module(name, path):
    """임의 경로의 Python 모듈을 로드한다. 이미 로드된 경우 캐시를 반환한다."""
    import sys
    if name in sys.modules:
        return sys.modules[name]
    spec = _ilu.spec_from_file_location(name, path)
    mod  = _ilu.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod
'''

# -----------------------------------------------------------------------
# 코드 분석
# -----------------------------------------------------------------------

def analyze_code_structure(file_paths):
    parts = []
    for path in file_paths:
        p = Path(path)
        if not p.exists() or p.suffix != ".py":
            continue
        try:
            tree = ast.parse(p.read_text(encoding="utf-8"))
        except SyntaxError:
            parts.append(f"[{p.name}] 파싱 실패")
            continue
        classes, functions, imports = [], [], []
        for node in ast.walk(tree):
            if isinstance(node, ast.ClassDef):
                methods = [n.name for n in ast.walk(node)
                           if isinstance(n, ast.FunctionDef)
                           and n.col_offset > node.col_offset]
                classes.append(f"  class {node.name}: [{', '.join(methods)}]")
            elif isinstance(node, ast.FunctionDef) and node.col_offset == 0:
                args = [a.arg for a in node.args.args]
                functions.append(f"  def {node.name}({', '.join(args)})")
            elif isinstance(node, ast.ImportFrom):
                imports.append(f"from {node.module} import ...")
            elif isinstance(node, ast.Import):
                for alias in node.names:
                    imports.append(f"import {alias.name}")
        section = [f"\n### {p.name}"]
        if imports:   section.append("imports: " + ", ".join(sorted(set(imports))))
        if classes:   section.append("classes:\n" + "\n".join(classes))
        if functions: section.append("top-level functions:\n" + "\n".join(functions))
        parts.append("\n".join(section))
    return "\n".join(parts) if parts else "(분석 없음)"


def read_files(file_paths):
    parts = []
    for path in file_paths:
        p = Path(path)
        if not p.exists():
            print(f"  [경고] 파일 없음: {path}")
            continue
        parts.append(f"### {p.name}\n```python\n{p.read_text(encoding='utf-8')}\n```")
    return "\n\n".join(parts)

# -----------------------------------------------------------------------
# Ollama
# -----------------------------------------------------------------------

def check_ollama(host, model):
    try:
        r = requests.get(f"{host}/api/tags", timeout=5)
        r.raise_for_status()
    except requests.exceptions.ConnectionError:
        print(f"\n[오류] Ollama 서버 연결 실패: {host}")
        print("먼저 실행하세요: ollama serve")
        sys.exit(1)
    models = [m["name"] for m in r.json().get("models", [])]
    if not any(model in m for m in models):
        print(f"\n[오류] '{model}' 모델 없음. 설치: ollama pull {model}")
        sys.exit(1)
    print(f"  ✓ Ollama 연결 확인 | 모델: {model}")


def call_ollama(prompt, model, host):
    print(f"  → Ollama 분석 중...")
    resp = requests.post(
        f"{host}/api/generate",
        json={"model": model,
              "prompt": f"{SYSTEM_PROMPT}\n\n{prompt}",
              "stream": False,
              "options": {"temperature": 0.1, "num_ctx": 16384}},
        timeout=1200,
    )
    resp.raise_for_status()
    raw = resp.json()["response"].strip()

    if "<think>" in raw:
        end = raw.find("</think>")
        if end != -1:
            raw = raw[end + len("</think>"):].strip()

    if "```json" in raw:
        raw = raw.split("```json")[1].split("```")[0]
    elif raw.startswith("```"):
        raw = "\n".join(raw.split("\n")[1:]).rstrip("`").strip()

    try:
        return json.loads(raw.strip())
    except json.JSONDecodeError as e:
        print(f"\n[오류] JSON 파싱 실패: {e}\n응답: {raw[:400]}")
        sys.exit(1)

# -----------------------------------------------------------------------
# 테스트 파일 조립
# -----------------------------------------------------------------------

def build_test_file(test_classes, source_files, fixtures_code, project_root, fixture_vars=None, config_vars=None):
    """
    [역할 분리]
    autotest.py  : import, 경로(PROJECT_ROOT), sys.modules mock, fixtures 삽입
    Ollama       : test_classes 의 메서드 body 만 생성
    fixtures.py  : 프로젝트별 헬퍼 함수 (_load_config, _make_logic 등)

    → 경로 오류가 구조적으로 불가능
    """
    root_str = str(project_root)

    # 입력 파일 절대경로 매핑
    file_map = {Path(f).stem: str(Path(f).resolve()) for f in source_files}

    header = (
        '"""\n'
        'generated_tests.py\n'
        f'AI 자동 생성 단위 테스트 — {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n'
        '인프라: autotest.py 직접 작성 / 픽스처: fixtures.py / 테스트 로직: Ollama 생성\n'
        '"""\n\n'
        'import sys, os, queue, importlib.util, pytest, time\n'
        'from unittest.mock import MagicMock, patch\n'
        'from threading import Event\n\n'
        f'# 프로젝트 루트 (autotest.py 위치 기준 — 자동 계산)\n'
        f'PROJECT_ROOT = r"{root_str}"\n'
        'sys.path.insert(0, PROJECT_ROOT)\n\n'
        '# 파일 경로 매핑 (autotest.py 자동 생성)\n'
        '_FILE_MAP = {\n'
    )
    for stem, abs_path in file_map.items():
        header += f'    "{stem}": r"{abs_path}",\n'
    header += '}\n\n'

    # fixtures.py 내용 삽입 (_FILE_MAP, MOCK_MODULES 정의 포함)
    # _FILE_MAP: dict = {} 선언만 제거 (런타임 덮어쓰기 방지)
    header += (
        '# ================================================================\n'
        '# 픽스처 (fixtures.py 내용 — 프로젝트별 커스터마이징 가능)\n'
        '# ================================================================\n\n'
    )
    fixtures_clean = "\n".join(
        line for line in fixtures_code.splitlines()
        if not line.strip().startswith("_FILE_MAP")
    )
    header += fixtures_clean.strip() + "\n\n"

    # MOCK_MODULES 루프 — fixtures.py 삽입 이후에 실행되므로 MOCK_MODULES 정의됨
    header += (
        '# 외부 의존성 Mock (MOCK_MODULES 는 fixtures.py 에서 정의)\n'
        'for _m in MOCK_MODULES:\n'
        '    sys.modules.setdefault(_m, MagicMock())\n\n'
    )

    # 테스트 클래스 조립
    class_blocks = []
    for tc in test_classes:
        cls_name = tc.get("class_name", "TestGenerated")
        desc     = tc.get("description", "")
        methods  = tc.get("methods", [])

        lines = [f"class {cls_name}:"]
        if desc:
            lines.append(f'    """{desc}"""')
        lines.append("")

        if not methods:
            lines.append("    pass")
        else:
            for m in methods:
                name = m.get("name", "test_unnamed")
                doc  = m.get("docstring", "")
                body = m.get("body", "pass")

                lines.append(f"    def {name}(self):")
                if doc:
                    lines.append(f'        """{doc}"""')

                # 픽스처 자동 주입 — cfg와 logic 동시에 필요한 경우도 처리
                # fixtures.py 의 FIXTURE_VARS / CONFIG_VARS 기준으로 주입
                # 특정 프로젝트 변수명을 autotest.py 가 알 필요 없음
                _fvars = fixture_vars or []
                _cvars = config_vars  or []
                needs_logic = any(v in body for v in _fvars)
                needs_cfg   = any(v in body for v in _cvars)

                if needs_logic:
                    lines.append("        logic, q_out, ST = _make_logic()")
                    if needs_cfg:
                        lines.append("        cfg = _load_config()")
                elif needs_cfg:
                    lines.append("        cfg = _load_config()")

                for bline in body.strip().split("\n"):
                    lines.append(f"        {bline.strip()}")
                lines.append("")

        class_blocks.append("\n".join(lines))

    return header + "\n\n".join(class_blocks) + "\n"

# -----------------------------------------------------------------------
# pytest 실행
# -----------------------------------------------------------------------

def run_pytest(test_code):
    OUTPUT_DIR.mkdir(exist_ok=True)
    test_file = OUTPUT_DIR / "generated_tests.py"
    test_file.write_text(test_code, encoding="utf-8")
    print(f"  → 테스트 파일 저장: {test_file}")

    result = subprocess.run(
        [sys.executable, "-m", "pytest", str(test_file),
         "-v", "--tb=short", "--no-header"],
        capture_output=True, text=True,
        encoding="utf-8", errors="replace",
    )

    stdout = result.stdout
    tests, fail_details, current_fail = [], {}, None

    for line in stdout.split("\n"):
        s = line.strip()
        for status in ("PASSED", "FAILED", "ERROR"):
            if status in s and "::" in s:
                parts = s.split("::")
                cls  = parts[1] if len(parts) > 2 else "기타"
                name = parts[-1].split(" ")[0]
                tests.append({"cls": cls, "name": name, "status": status})
                if status in ("FAILED", "ERROR"):
                    current_fail = name
                    fail_details[name] = []
                break
        else:
            if current_fail and s and not s.startswith(("_", "=")):
                fail_details[current_fail].append(s)

    summary_line = next(
        (l.strip() for l in reversed(stdout.split("\n"))
         if "passed" in l or "failed" in l or "error" in l), ""
    )

    passed = sum(1 for t in tests if t["status"] == "PASSED")
    failed = sum(1 for t in tests if t["status"] in ("FAILED", "ERROR"))
    total  = len(tests)

    ERROR_KW = ("ImportError", "ModuleNotFoundError", "SyntaxError",
                "FileNotFoundError", "collected 0 items")
    collection_error = (
        result.returncode in (2, 3, 4, 5)
        or (total == 0 and result.returncode != 0)
        or any(kw in stdout for kw in ERROR_KW)
    )
    cerr_msg = ""
    if collection_error:
        for line in stdout.split("\n"):
            if any(kw in line for kw in ("Error", "ERRORS", "Exception")):
                cerr_msg += line.strip() + "\n"
        cerr_msg = cerr_msg.strip() or stdout[:400]

    return {
        "return_code":          result.returncode,
        "summary_line":         summary_line,
        "tests":                tests,
        "fail_details":         fail_details,
        "stdout":               stdout,
        "total":                total,
        "passed":               passed,
        "failed":               failed,
        "pass_rate":            f"{passed / total * 100:.1f}%" if total else "N/A",
        "collection_error":     collection_error,
        "collection_error_msg": cerr_msg,
    }

# -----------------------------------------------------------------------
# 산출물 저장
# -----------------------------------------------------------------------

def save_report(summary, model, files, pr):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if pr["collection_error"]:
        status = "💥  SETUP ERROR"
    elif pr["failed"] > 0:
        status = "⚠️  BUG FOUND"
    else:
        status = "✅ ALL PASS"

    lines = [
        "# 단위 테스트 자동화 리포트", "",
        "| 항목 | 내용 |", "|------|------|",
        f"| 생성 도구 | Ollama ({model}) |",
        f"| 실행 일시 | {now} |",
        f"| 대상 파일 | {', '.join(Path(f).name for f in files)} |",
        f"| 수정 요약 | {summary} |",
        f"| 최종 결과 | {status} |",
        f"| 통과율 | {pr['pass_rate']} ({pr['passed']}/{pr['total']}) |",
        "", "---", "",
    ]

    if pr["collection_error"]:
        lines += ["## 💥 셋업 오류", "", "```"]
        lines += pr["collection_error_msg"].split("\n")[:15]
        lines += ["```", ""]

    passed_tests = [t for t in pr["tests"] if t["status"] == "PASSED"]
    if passed_tests:
        lines.append(f"## ✅ 통과 ({len(passed_tests)}개)"); lines.append("")
        groups = {}
        for t in passed_tests:
            groups.setdefault(t["cls"], []).append(t["name"])
        for cls, names in groups.items():
            lines.append(f"**{cls}**")
            for n in names: lines.append(f"- ✅ `{n}`")
            lines.append("")

    failed_tests = [t for t in pr["tests"] if t["status"] in ("FAILED", "ERROR")]
    if failed_tests:
        lines += ["---", "", f"## ❌ 실패 ({len(failed_tests)}개)", "",
                  "> 코드 수정이 필요한 결함입니다.", ""]
        for t in failed_tests:
            detail = pr["fail_details"].get(t["name"], [])
            lines.append(f"### ❌ `{t['name']}`")
            if detail: lines += ["```"] + detail[:6] + ["```"]
            lines.append("")

    lines += ["---", "", "## pytest 전체 출력", "", "```",
              pr["stdout"].strip(), "```"]
    path = OUTPUT_DIR / "test_result_report.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def save_json(summary, model, files, pr, ui_count):
    data = {
        "tool": f"Ollama ({model})",
        "timestamp": datetime.now().isoformat(),
        "summary": summary,
        "files": [Path(f).name for f in files],
        "result": ("COLLECTION_ERROR" if pr["collection_error"]
                   else "FAIL" if pr["failed"] > 0 else "PASS"),
        "pytest": {
            "total": pr["total"], "passed": pr["passed"],
            "failed": pr["failed"], "pass_rate": pr["pass_rate"],
            "collection_error": pr["collection_error"],
        },
        "failed_tests":  [t["name"] for t in pr["tests"]
                          if t["status"] in ("FAILED", "ERROR")],
        "ui_test_count": ui_count,
    }
    path = OUTPUT_DIR / "test_result.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def save_excel(ui_cases, summary, model):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [경고] openpyxl 없음 — pip install openpyxl")
        return None

    C = {"navy":"1E2761","ice":"C8D8F8","white":"FFFFFF",
         "gray":"F4F6FB","red":"FCE4E4","yellow":"FFF9CC","green":"D6F0E0"}
    PRI = {"P1":C["red"], "P2":C["yellow"], "P3":C["green"]}

    def _f(bold=False, sz=10, color="1E2761"):
        return Font(name="Arial", bold=bold, size=sz, color=color)
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _bd():
        s = Side(border_style="thin", color="C0C8D8")
        return Border(left=s, right=s, top=s, bottom=s)
    def _al(h="left", wrap=True):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    wb = Workbook(); ws = wb.active
    ws.title = "UI 수동 테스트"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws.merge_cells("A1:J1")
    ws["A1"] = f"UI 수동 테스트  |  {summary}"
    ws["A1"].font = _f(bold=True, sz=13, color=C["white"])
    ws["A1"].fill = _fill(C["navy"])
    ws["A1"].alignment = _al("center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = (f"생성 도구: Ollama ({model})  |  "
                f"생성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws["A2"].font = _f(sz=9, color="555555")
    ws["A2"].fill = _fill(C["ice"])
    ws["A2"].alignment = _al("center")
    ws.row_dimensions[2].height = 16

    ws.merge_cells("A3:J3")
    ws["A3"] = "※ 코드로 검증 불가능한 UI/사용자 시나리오입니다. 테스터가 Pass/Fail을 기입하세요."
    ws["A3"].font = _f(sz=9, color="7A5500", bold=True)
    ws["A3"].fill = _fill("FFFBE6")
    ws["A3"].alignment = _al("left")
    ws.row_dimensions[3].height = 15

    COLS = [("NO",5),("구분",14),("테스트 항목",28),("사전 조건",22),
            ("테스트 절차",35),("기대 결과",28),("실제 결과",18),
            ("Pass/Fail",11),("비고",14),("우선순위",10)]
    for col, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = _f(bold=True, sz=10, color=C["white"])
        c.fill = _fill(C["navy"])
        c.alignment = _al("center")
        c.border = _bd()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    for ri, case in enumerate(ui_cases, 5):
        bg = C["gray"] if ri % 2 == 0 else C["white"]
        row = [case.get("no", ri-4), case.get("category",""),
               case.get("title",""), case.get("precondition",""),
               case.get("steps",""), case.get("expected",""),
               "", "", "", case.get("priority","P2")]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.font = _f(sz=10)
            cell.border = _bd()
            cell.alignment = _al("center" if ci in (1,8,10) else "left")
        pc = ws.cell(row=ri, column=10)
        pc.fill = _fill(PRI.get(case.get("priority","P2"), C["white"]))
        pc.font = _f(bold=True, sz=10, color=C["navy"])
        ws.row_dimensions[ri].height = 50

    last = len(ui_cases) + 6
    ws.merge_cells(f"A{last}:B{last}")
    ws.cell(row=last, column=1, value="우선순위 범례").fill = _fill(C["ice"])
    for i, (p, c, d) in enumerate([
        ("P1",C["red"],"필수"),("P2",C["yellow"],"중요"),("P3",C["green"],"권장")], 1):
        r = last + i
        ws.cell(row=r, column=1, value=p).fill = _fill(c)
        ws.cell(row=r, column=1).font = _f(bold=True)
        ws.cell(row=r, column=1).alignment = _al("center")
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=d).font = _f(sz=9)

    path = OUTPUT_DIR / "UI_테스트목록.xlsx"
    wb.save(path)
    return path

# -----------------------------------------------------------------------
# 메인 파이프라인
# -----------------------------------------------------------------------

def run_pipeline(summary, files, model, host):
    OUTPUT_DIR.mkdir(exist_ok=True)
    project_root = Path(__file__).resolve().parent

    print(f"\n{'='*58}")
    print(f"  자동 단위 테스트 파이프라인  (Ollama 전용)")
    print(f"{'='*58}")
    print(f"  모델        : {model}")
    print(f"  요약        : {summary}")
    print(f"  파일        : {', '.join(Path(f).name for f in files)}")
    print(f"  프로젝트 루트: {project_root}")
    print(f"{'='*58}\n")

    print("[0/4] Ollama 서버 확인 중...")
    check_ollama(host, model)

    print("\n[1/4] 코드 구조 분석 중...")
    code_structure = analyze_code_structure(files)
    files_content  = read_files(files)
    print(textwrap.indent(code_structure, "      "))

    print("\n[2/4] Ollama에게 테스트 로직 생성 요청 중...")
    prompt = USER_PROMPT_TEMPLATE.format(
        summary=summary,
        code_structure=code_structure,
        files_content=files_content,
    )
    ai_result    = call_ollama(prompt, model, host)
    test_classes = ai_result.get("test_classes", [])
    ui_cases     = ai_result.get("ui_test_cases", [])
    method_count = sum(len(tc.get("methods", [])) for tc in test_classes)
    print(f"      생성 완료 — 클래스 {len(test_classes)}개 / "
          f"테스트 {method_count}개 / UI 케이스 {len(ui_cases)}개")

    print("\n[3/4] 테스트 파일 조립 및 pytest 실행 중...")
    fixtures_code = load_fixtures(project_root)
    # fixtures.py 에서 FIXTURE_VARS, CONFIG_VARS, 호출식 추출
    fx_vars   = _extract_var(fixtures_code, 'FIXTURE_VARS', [])
    cfg_vars  = _extract_var(fixtures_code, 'CONFIG_VARS',  ['cfg'])
    mk_call   = _extract_var(fixtures_code, 'MAKE_LOGIC_CALL',   None)
    ld_call   = _extract_var(fixtures_code, 'LOAD_CONFIG_CALL',  None)
    test_code = build_test_file(
        test_classes, files, fixtures_code, project_root,
        fixture_vars=fx_vars, config_vars=cfg_vars,
    )
    pr = run_pytest(test_code)

    if pr["collection_error"]:
        print(f"      💥 셋업 오류 (exit {pr['return_code']})")
        for ln in pr["collection_error_msg"].split("\n")[:3]:
            if ln.strip(): print(f"         {ln.strip()[:110]}")
    else:
        icon = "✅" if pr["failed"] == 0 else "⚠️ "
        print(f"      {icon} {pr['summary_line']}")
        if pr["failed"] > 0:
            names = [t["name"] for t in pr["tests"]
                     if t["status"] in ("FAILED","ERROR")]
            print(f"      실패: {', '.join(names)}")

    print("\n[4/4] 산출물 생성 중...")
    save_report(summary, model, files, pr)
    save_json(summary, model, files, pr, len(ui_cases))
    e_path = save_excel(ui_cases, summary, model)

    print(f"\n{'='*58}")
    if pr["collection_error"]:
        print("  ⚠️  셋업 오류 — test_result_report.md 확인")
    else:
        print("  완료!")
    print(f"  ├── output/generated_tests.py    ({pr['total']}개 테스트)")
    print(f"  ├── output/test_result_report.md ({pr['passed']}✅  {pr['failed']}❌)")
    print(f"  ├── output/test_result.json")
    if e_path:
        print(f"  └── output/UI_테스트목록.xlsx   ({len(ui_cases)}개 케이스)")
    print(f"{'='*58}\n")
    return pr


def main():
    parser = argparse.ArgumentParser(
        description="범용 Python 단위 테스트 자동화 — Ollama 전용",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            예시:
              python autotest.py --summary "로그인 기능 추가" --files config/config.py logic/LogicProcess.py
              python autotest.py --summary "결제 리팩토링" --files pay.py --model qwen3-coder:8b
        """),
    )
    parser.add_argument("--summary", required=True, help="수정 내용 요약")
    parser.add_argument("--files",   required=True, nargs="+",
                        help="수정된 Python 파일 경로")
    parser.add_argument("--model",   default=DEFAULT_MODEL)
    parser.add_argument("--host",    default=DEFAULT_HOST)
    args = parser.parse_args()
    pr = run_pipeline(args.summary, args.files, args.model, args.host)
    sys.exit(0 if (pr["failed"] == 0 and not pr["collection_error"]) else 1)


if __name__ == "__main__":
    main()